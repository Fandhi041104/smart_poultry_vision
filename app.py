from flask import Flask, render_template, Response, jsonify, request, send_file
import cv2
import mysql.connector
from mysql.connector import Error
import datetime
import os
from ultralytics import YOLO
import threading
import time
import numpy as np

app = Flask(__name__)

MODEL_PATH       = "models/best.pt"
SNAPSHOT_FOLDER  = "static/snapshots"
RECORDING_FOLDER = "static/recordings"
VIDEO_FOLDER     = "test_videos"

# "video"  -> tes pakai file video
# "camera" -> IP Camera asli (fallback otomatis ke webcam kalau RTSP gagal)
# "webcam" -> tes cepat pakai webcam laptop saja (tidak coba RTSP sama sekali)
OPERATION_MODE = "video"
VIDEO_LOOP     = True

CAMERA_CONFIG = {
    1: {'ip': '192.168.0.14',  'user': 'fandhi', 'pass': 'tugas_akhir',
        'video_path': 'test_videos/test_broiler4.mp4'},
    2: {'ip': '192.168.0.12', 'user': 'fandhi', 'pass': 'fandhi_ta',
        'video_path': 'test_videos/test_broiler3.mp4'}
}

CONFIDENCE_THRESHOLD = 0.45
DETECTION_COOLDOWN   = 2
MIN_BOX_AREA         = 1500
MAX_BOX_AREA         = 800000
IOU_THRESHOLD        = 0.3
DETECTION_FRAME_SKIP = 8
RESIZE_WIDTH         = 640
JPEG_QUALITY         = 78
INFERENCE_SIZE       = 640
SNAPSHOT_SIZE        = 640
MAX_SNAPSHOTS        = 10

os.makedirs(SNAPSHOT_FOLDER, exist_ok=True)
os.makedirs(RECORDING_FOLDER, exist_ok=True)
os.makedirs("models", exist_ok=True)
os.makedirs(VIDEO_FOLDER, exist_ok=True)

cameras              = {1: None, 2: None}
last_detection_time  = {1: 0, 2: 0}
last_frame_wh        = {1: (RESIZE_WIDTH, 360), 2: (RESIZE_WIDTH, 360)}
model                = None
model_lock           = threading.Lock()

# ── Rekaman, terpisah per kamera ────────────────────────────────
recording_state = {
    1: {'is_recording': False, 'writer': None, 'filename': '', 'start': None},
    2: {'is_recording': False, 'writer': None, 'filename': '', 'start': None}
}
recording_locks = {1: threading.Lock(), 2: threading.Lock()}

# ── Bounding box terakhir (skala display) khusus untuk overlay rekaman ──
display_detections = {1: [], 2: []}
display_lock        = threading.Lock()

latest_detection = {
    'status': 'Tidak Ada Deteksi', 'confidence': 0,
    'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    'disease_counts': {'Sehat': 0, 'Coccidiosis': 0, 'Newcastle': 0, 'Salmonella': 0},
    'total_detected': 0, 'camera': 0
}
latest_lock = threading.Lock()

detection_snapshots = []
snapshot_lock       = threading.Lock()

CLASS_NAMES = {0: 'Sehat', 1: 'Coccidiosis', 2: 'Newcastle', 3: 'Salmonella'}
COLOR_MAP = {'Sehat': (0,255,0), 'Coccidiosis': (0,165,255), 'Newcastle': (0,0,255), 'Salmonella': (255,0,255)}

MYSQL_CONFIG = {'host':'localhost','user':'root','password':'','database':'poultry_detection','port':3306,'autocommit':True}


def get_db_connection():
    try:
        return mysql.connector.connect(**MYSQL_CONFIG)
    except Error as e:
        print(f"Error koneksi MySQL: {e}")
        return None


def init_db():
    conn = get_db_connection()
    print("Terhubung ke MySQL" if conn else "Gagal terhubung ke MySQL")
    if conn: conn.close()


init_db()


def save_detection(disease, confidence, snapshot_filename=''):
    try:
        conn = get_db_connection()
        if not conn: return
        cursor = conn.cursor()
        ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("INSERT INTO detections (disease, confidence, timestamp, image_path) VALUES (%s,%s,%s,%s)",
                       (disease, confidence, ts, snapshot_filename))
        cursor.close(); conn.close()
    except Error as e:
        print(f"MySQL Error: {e}")


def load_model():
    global model
    try:
        if os.path.exists(MODEL_PATH):
            print(f"Memuat model dari {MODEL_PATH}")
            model = YOLO(MODEL_PATH)
            _ = model(np.zeros((640,640,3), dtype=np.uint8), verbose=False)
            print("Model berhasil dimuat")
        else:
            print("File model tidak ditemukan")
    except Exception as e:
        print(f"Error memuat model: {e}")


load_model()


def connect_camera(camera_id):
    global cameras
    cfg = CAMERA_CONFIG[camera_id]
    webcam_index = camera_id - 1  # kamera 1 -> webcam 0, kamera 2 -> webcam 1
    try:
        if OPERATION_MODE == "webcam":
            print(f"[Kamera {camera_id}] Mencoba webcam index {webcam_index}...")
            cap = cv2.VideoCapture(webcam_index)
            if cap.isOpened():
                cameras[camera_id] = cap
                print(f"[Kamera {camera_id}] Webcam terhubung")
                return True
            print(f"[Kamera {camera_id}] Webcam index {webcam_index} tidak tersedia")
            return False

        if OPERATION_MODE == "video":
            if not os.path.exists(cfg['video_path']):
                print(f"[Kamera {camera_id}] File video tidak ditemukan: {cfg['video_path']}")
                return False
            print(f"[Kamera {camera_id}] Memuat video: {cfg['video_path']}")
            cap = cv2.VideoCapture(cfg['video_path'])
        else:  # mode "camera" -> IP Camera asli
            rtsp_url = f"rtsp://{cfg['user']}:{cfg['pass']}@{cfg['ip']}:554/stream1"
            print(f"[Kamera {camera_id}] Menghubungkan ke {cfg['ip']}...")
            cap = cv2.VideoCapture(rtsp_url, cv2.CAP_FFMPEG)
            cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
            cap.set(cv2.CAP_PROP_FPS, 30)
            cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)

        if cap.isOpened():
            cameras[camera_id] = cap
            print(f"[Kamera {camera_id}] Terhubung")
            return True

        print(f"[Kamera {camera_id}] Gagal terhubung")
        if OPERATION_MODE == "camera":
            print(f"[Kamera {camera_id}] Mencoba fallback webcam index {webcam_index}...")
            cap = cv2.VideoCapture(webcam_index)
            if cap.isOpened():
                cameras[camera_id] = cap
                print(f"[Kamera {camera_id}] Fallback webcam berhasil")
                return True
        return False
    except Exception as e:
        print(f"[Kamera {camera_id}] Error: {e}")
        return False


def create_detection_snapshot(frame, detections_data, camera_id):
    if not detections_data:
        return None
    snapshot = frame.copy()
    for det in detections_data:
        box, disease, conf = det['box'], det['disease'], det['confidence']
        x1, y1, x2, y2 = map(int, box.xyxy[0])
        color = COLOR_MAP.get(disease, (255,255,255))
        cv2.rectangle(snapshot, (x1,y1), (x2,y2), color, 4)
        label = f"{disease}: {conf:.1f}%"
        (w,h), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 1.0, 2)
        cv2.rectangle(snapshot, (x1,y1-h-15), (x1+w+10,y1), color, -1)
        cv2.putText(snapshot, label, (x1+5,y1-8), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255,255,255), 2)

    ht, wd = snapshot.shape[:2]
    if wd > SNAPSHOT_SIZE:
        scale = SNAPSHOT_SIZE / wd
        snapshot = cv2.resize(snapshot, (int(wd*scale), int(ht*scale)))

    info = f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Kamera {camera_id} | {len(detections_data)} objek"
    cv2.putText(snapshot, info, (10, snapshot.shape[0]-15), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 2)

    filename = f"cam{camera_id}_detection_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
    cv2.imwrite(os.path.join(SNAPSHOT_FOLDER, filename), snapshot, [cv2.IMWRITE_JPEG_QUALITY, 95])
    return filename


def start_video_writer(camera_id):
    state = recording_state[camera_id]
    w, h = last_frame_wh[camera_id]
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"cam{camera_id}_rekaman_{ts}.mp4"
    filepath = os.path.join(RECORDING_FOLDER, filename)
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    state['writer']   = cv2.VideoWriter(filepath, fourcc, 15.0, (w, h))
    state['filename'] = filename
    state['start']    = datetime.datetime.now()
    print(f"[Kamera {camera_id}] Rekaman dimulai: {filepath}")


def stop_video_writer(camera_id):
    state = recording_state[camera_id]
    with recording_locks[camera_id]:
        if state['writer']:
            state['writer'].release()
            state['writer'] = None
        state['is_recording'] = False
    print(f"[Kamera {camera_id}] Rekaman dihentikan: {state['filename']}")


def generate_frames(camera_id):
    global cameras, latest_detection, last_detection_time, detection_snapshots, last_frame_wh

    if cameras.get(camera_id) is None or not cameras[camera_id].isOpened():
        connect_camera(camera_id)

    frame_count  = 0
    local_status = "Tidak Ada Deteksi"
    local_conf   = 0
    print(f"[Kamera {camera_id}] Stream dimulai")

    while True:
        cap = cameras.get(camera_id)
        if cap is None or not cap.isOpened():
            time.sleep(0.5)
            connect_camera(camera_id)
            continue

        success, frame = cap.read()
        if not success:
            if OPERATION_MODE == "video" and VIDEO_LOOP:
                cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
                continue
            print(f"[Kamera {camera_id}] Menghubungkan ulang...")
            time.sleep(0.1)
            connect_camera(camera_id)
            continue

        frame_count    += 1
        current_time    = time.time()
        original_frame  = frame.copy()
        h_orig, w_orig  = frame.shape[:2]
        scale_display   = RESIZE_WIDTH / w_orig if w_orig > RESIZE_WIDTH else 1.0

        if frame_count % DETECTION_FRAME_SKIP == 0 and model is not None:
            try:
                with model_lock:
                    results = model(original_frame, conf=CONFIDENCE_THRESHOLD, iou=IOU_THRESHOLD,
                                    verbose=False, imgsz=INFERENCE_SIZE, max_det=20)

                if len(results[0].boxes) > 0:
                    detections_data = []
                    disease_counts  = {'Sehat':0,'Coccidiosis':0,'Newcastle':0,'Salmonella':0}

                    for box in results[0].boxes:
                        conf = float(box.conf[0])
                        x1,y1,x2,y2 = box.xyxy[0]
                        area = (x2-x1)*(y2-y1)
                        if conf >= CONFIDENCE_THRESHOLD and MIN_BOX_AREA <= area <= MAX_BOX_AREA:
                            cls_id  = int(box.cls[0])
                            disease = CLASS_NAMES.get(cls_id, f'Unknown_{cls_id}')
                            detections_data.append({'box':box,'disease':disease,'confidence':conf*100})
                            disease_counts[disease] = disease_counts.get(disease,0)+1

                    if detections_data:
                        total        = len(detections_data)
                        main_disease = max(disease_counts, key=disease_counts.get)
                        main_count   = disease_counts[main_disease]
                        max_conf     = round(max(d['confidence'] for d in detections_data), 2)
                        now_str      = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                        local_status = f"{main_disease} ({main_count}x)" if total>1 else main_disease
                        local_conf   = max_conf

                        with display_lock:
                            boxes = []
                            for det in detections_data:
                                bx1,by1,bx2,by2 = map(float, det['box'].xyxy[0])
                                boxes.append({
                                    'x1': int(bx1*scale_display), 'y1': int(by1*scale_display),
                                    'x2': int(bx2*scale_display), 'y2': int(by2*scale_display),
                                    'disease': det['disease'], 'confidence': det['confidence'],
                                    'color': COLOR_MAP.get(det['disease'], (255,255,255))
                                })
                            display_detections[camera_id] = boxes

                        with latest_lock:
                            latest_detection = {'status':local_status,'confidence':max_conf,'timestamp':now_str,
                                                'disease_counts':disease_counts,'total_detected':total,'camera':camera_id}

                        if current_time - last_detection_time[camera_id] >= DETECTION_COOLDOWN:
                            snap_name = create_detection_snapshot(original_frame, detections_data, camera_id)
                            if snap_name:
                                with snapshot_lock:
                                    detection_snapshots.insert(0, {
                                        'filename':snap_name,'url':f'/static/snapshots/{snap_name}',
                                        'disease':main_disease,'confidence':max_conf,'timestamp':now_str,
                                        'counts':disease_counts,'camera':camera_id})
                                    while len(detection_snapshots) > MAX_SNAPSHOTS:
                                        old = detection_snapshots.pop()
                                        old_path = os.path.join(SNAPSHOT_FOLDER, old['filename'])
                                        if os.path.exists(old_path): os.remove(old_path)
                            for det in detections_data:
                                threading.Thread(target=save_detection,
                                    args=(det['disease'], det['confidence']/100, snap_name or ''), daemon=True).start()
                            last_detection_time[camera_id] = current_time
                            print(f"[Kamera {camera_id}] Disimpan: {total} objek - {disease_counts}")
            except Exception as e:
                print(f"[Kamera {camera_id}] Error deteksi: {e}")

        if w_orig > RESIZE_WIDTH:
            frame = cv2.resize(frame, None, fx=scale_display, fy=scale_display)

        h_disp, w_disp = frame.shape[:2]
        last_frame_wh[camera_id] = (w_disp, h_disp)

        # ── Overlay teks (tanpa bounding box) ──
        cv2.putText(frame, f"Kam{camera_id} | Threshold {CONFIDENCE_THRESHOLD*100:.0f}%",
                    (8,22), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255,255,255), 2)
        if local_conf > 0:
            cv2.putText(frame, f"{local_status} - {local_conf:.1f}%", (8,43),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0,255,0), 2)

        # ── Indikator & tulis rekaman (dengan bounding box) ──
        rec_state = recording_state[camera_id]
        if rec_state['is_recording']:
            secs = int((datetime.datetime.now()-rec_state['start']).total_seconds()) if rec_state['start'] else 0
            elapsed = f"{secs//3600:02d}:{(secs%3600)//60:02d}:{secs%60:02d}"
            cv2.circle(frame, (frame.shape[1]-20,18), 7, (0,0,255), -1)
            cv2.putText(frame, f"REC {elapsed}", (frame.shape[1]-130,24),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0,0,255), 2)

            rec_frame = frame.copy()
            with display_lock:
                for det in display_detections[camera_id]:
                    color = det['color']
                    cv2.rectangle(rec_frame, (det['x1'],det['y1']), (det['x2'],det['y2']), color, 3)
                    label = f"{det['disease']}: {det['confidence']:.1f}%"
                    (tw,th),_ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.55, 2)
                    cv2.rectangle(rec_frame, (det['x1'],det['y1']-th-10), (det['x1']+tw+8,det['y1']), color, -1)
                    cv2.putText(rec_frame, label, (det['x1']+4,det['y1']-4),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255,255,255), 2)

            with recording_locks[camera_id]:
                if rec_state['writer'] and rec_state['writer'].isOpened():
                    rec_state['writer'].write(rec_frame)

        ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])
        if ret:
            yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')

        if OPERATION_MODE == "video":
            time.sleep(0.033)


# ════════════════════════════════════════════
#  ROUTES
# ════════════════════════════════════════════
@app.route('/')
def index(): return render_template('dashboard.html')

@app.route('/dashboard')
def dashboard(): return render_template('dashboard.html')

@app.route('/riwayat')
def riwayat(): return render_template('riwayat.html')

@app.route('/edukasi')
def edukasi(): return render_template('edukasi.html')

@app.route('/video_feed/<int:camera_id>')
def video_feed(camera_id):
    if camera_id not in CAMERA_CONFIG:
        return "Kamera tidak ditemukan", 404
    return Response(generate_frames(camera_id), mimetype='multipart/x-mixed-replace; boundary=frame')


# ════════════════════════════════════════════
#  API — Deteksi
# ════════════════════════════════════════════
@app.route('/api/latest_detection')
def get_latest_detection():
    with latest_lock:
        return jsonify(latest_detection)

@app.route('/api/detection_snapshots')
def get_detection_snapshots():
    with snapshot_lock:
        return jsonify({'snapshots': list(detection_snapshots)})

@app.route('/api/camera_status')
def camera_status():
    status, online = {}, 0
    for cid, cap in cameras.items():
        ok = bool(cap and cap.isOpened())
        status[f'camera{cid}'] = 'online' if ok else 'offline'
        online += 1 if ok else 0
    status['summary'] = f"{online}/{len(cameras)} Terhubung"
    return jsonify(status)

@app.route('/api/model_info')
def model_info():
    if model:
        return jsonify({'loaded':True,'confidence':CONFIDENCE_THRESHOLD,'inference_size':INFERENCE_SIZE})
    return jsonify({'loaded': False})


# ════════════════════════════════════════════
#  API — Database (statistik, riwayat, export)
# ════════════════════════════════════════════
@app.route('/api/statistics')
def get_statistics():
    try:
        conn = get_db_connection()
        if not conn: return jsonify({'error':'Koneksi database gagal'}), 500
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT COUNT(*) as total FROM detections")
        total = cursor.fetchone()['total']
        cursor.execute("SELECT disease, COUNT(*) as count FROM detections GROUP BY disease")
        by_disease = {r['disease']: r['count'] for r in cursor.fetchall()}
        cursor.execute("SELECT disease, confidence, timestamp FROM detections ORDER BY id DESC LIMIT 10")
        recent = [{'disease':r['disease'],'confidence':round(r['confidence']*100,2),
                   'timestamp':r['timestamp'].strftime('%Y-%m-%d %H:%M:%S')} for r in cursor.fetchall()]
        cursor.close(); conn.close()
        return jsonify({'total':total,'by_disease':by_disease,'recent':recent})
    except Error as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/statistics_today')
def get_statistics_today():
    try:
        conn = get_db_connection()
        if not conn: return jsonify({'error':'Koneksi database gagal'}), 500
        cursor = conn.cursor(dictionary=True)
        today = datetime.date.today().strftime('%Y-%m-%d')
        cursor.execute("SELECT disease, COUNT(*) as count FROM detections WHERE DATE(timestamp)=%s GROUP BY disease", (today,))
        today_data = {r['disease']: r['count'] for r in cursor.fetchall()}
        cursor.execute("SELECT COUNT(*) as total FROM detections")
        total_all = cursor.fetchone()['total']
        cursor.close(); conn.close()
        return jsonify({'today':today_data,'total_all':total_all,'date':today})
    except Error as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/history_all')
def get_history_all():
    try:
        conn = get_db_connection()
        if not conn: return jsonify({'error':'Koneksi database gagal'}), 500
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT disease, confidence, timestamp, image_path FROM detections ORDER BY id DESC")
        history = []
        for r in cursor.fetchall():
            snap_url = f"/static/snapshots/{r['image_path']}" if r['image_path'] else ''
            history.append({'disease':r['disease'],'confidence':round(r['confidence']*100,2),
                            'timestamp':r['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),'snapshot_url':snap_url})
        cursor.close(); conn.close()
        return jsonify(history)
    except Error as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export_history')
def export_history():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        disease, from_date, to_date = request.args.get('disease',''), request.args.get('from',''), request.args.get('to','')
        conn = get_db_connection()
        if not conn: return jsonify({'error':'Koneksi database gagal'}), 500
        cursor = conn.cursor(dictionary=True)
        query, params = "SELECT disease, confidence, timestamp FROM detections WHERE 1=1", []
        if disease:   query += " AND disease=%s"; params.append(disease)
        if from_date: query += " AND DATE(timestamp)>=%s"; params.append(from_date)
        if to_date:   query += " AND DATE(timestamp)<=%s"; params.append(to_date)
        query += " ORDER BY id DESC"
        cursor.execute(query, params)
        rows = cursor.fetchall()
        cursor.close(); conn.close()

        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Riwayat Deteksi"
        headers = ['No','Penyakit','Tingkat Keyakinan (%)','Tanggal','Waktu']
        hfill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        hfont = Font(color="FFFFFF", bold=True)
        for col,h in enumerate(headers,1):
            c = ws.cell(row=1,column=col,value=h); c.fill=hfill; c.font=hfont
            c.alignment = Alignment(horizontal='center')
        for idx,r in enumerate(rows,1):
            ts = r['timestamp']
            d_ = ts.strftime('%Y-%m-%d') if hasattr(ts,'strftime') else str(ts).split(' ')[0]
            t_ = ts.strftime('%H:%M:%S') if hasattr(ts,'strftime') else str(ts).split(' ')[1]
            ws.cell(row=idx+1,column=1,value=idx)
            ws.cell(row=idx+1,column=2,value=r['disease'])
            ws.cell(row=idx+1,column=3,value=round(r['confidence']*100,2))
            ws.cell(row=idx+1,column=4,value=d_)
            ws.cell(row=idx+1,column=5,value=t_)
        for col,w in zip(['A','B','C','D','E'],[6,16,22,14,12]):
            ws.column_dimensions[col].width = w
        suffix = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        if disease: suffix = f"{disease}_{suffix}"
        output = BytesIO(); wb.save(output); output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'riwayat_deteksi_{suffix}.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/snapshots/all')
def get_all_snapshots():
    try:
        page, per_page = int(request.args.get('page',1)), int(request.args.get('per_page',12))
        files = []
        if os.path.exists(SNAPSHOT_FOLDER):
            for f in os.listdir(SNAPSHOT_FOLDER):
                if f.lower().endswith(('.jpg','.jpeg','.png')):
                    timestamp, camera_label, sort_key = '', '', ''
                    try:
                        parts = f.replace('.jpg','').split('_')
                        if parts[0].startswith('cam'):
                            camera_label = parts[0].replace('cam','Kamera ')
                            parts = parts[1:]
                        if len(parts) >= 3:
                            d,t = parts[1], parts[2]
                            timestamp = f"{d[:4]}-{d[4:6]}-{d[6:8]} {t[:2]}:{t[2:4]}:{t[4:6]}"
                            sort_key = d + t  # YYYYMMDDHHMMSS, dipakai untuk urutkan murni berdasar waktu
                    except: pass
                    files.append({'filename':f,'url':f'/static/snapshots/{f}','timestamp':timestamp,
                                  'camera':camera_label,'_sort_key':sort_key})

        # Urutkan berdasarkan waktu sebenarnya (terbaru di atas), bukan nama file mentah,
        # supaya hasil dari kamera 1 dan kamera 2 tetap terselip sesuai urutan waktu deteksi
        files.sort(key=lambda x: x['_sort_key'], reverse=True)
        for f in files:
            f.pop('_sort_key', None)

        total = len(files)
        total_pages = max(1,(total+per_page-1)//per_page)
        paged = files[(page-1)*per_page:page*per_page]
        return jsonify({'snapshots':paged,'total':total,'page':page,'total_pages':total_pages,'per_page':per_page})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/set_threshold', methods=['POST'])
def set_threshold():
    global CONFIDENCE_THRESHOLD
    try:
        val = float(request.get_json().get('threshold', 0.45))
        if 0.1 <= val <= 1.0:
            CONFIDENCE_THRESHOLD = val
            return jsonify({'success':True,'threshold':CONFIDENCE_THRESHOLD})
        return jsonify({'success':False,'error':'Nilai harus 0.1–1.0'}), 400
    except Exception as e:
        return jsonify({'success':False,'error':str(e)}), 500


# ════════════════════════════════════════════
#  API — Rekaman (per kamera)
# ════════════════════════════════════════════
@app.route('/api/recording/start/<int:camera_id>', methods=['POST'])
def start_recording(camera_id):
    if camera_id not in recording_state:
        return jsonify({'success':False,'message':'Kamera tidak ditemukan'}), 404
    state = recording_state[camera_id]
    if state['is_recording']:
        return jsonify({'success':False,'message':'Rekaman sudah berjalan'})
    cap = cameras.get(camera_id)
    if cap is None or not cap.isOpened():
        return jsonify({'success':False,'message':'Kamera/video tidak terhubung'})
    with recording_locks[camera_id]:
        start_video_writer(camera_id)
        state['is_recording'] = True
    return jsonify({'success':True,'message':'Rekaman dimulai','filename':state['filename']})

@app.route('/api/recording/stop/<int:camera_id>', methods=['POST'])
def stop_recording(camera_id):
    if camera_id not in recording_state:
        return jsonify({'success':False,'message':'Kamera tidak ditemukan'}), 404
    state = recording_state[camera_id]
    if not state['is_recording']:
        return jsonify({'success':False,'message':'Tidak ada rekaman yang berjalan'})
    saved = state['filename']
    stop_video_writer(camera_id)
    return jsonify({'success':True,'message':'Rekaman berhasil disimpan','filename':saved,
                    'download_url':f'/api/recording/download/{saved}'})

@app.route('/api/recording/status/<int:camera_id>')
def recording_status(camera_id):
    state = recording_state.get(camera_id, {})
    elapsed = 0
    if state.get('is_recording') and state.get('start'):
        elapsed = int((datetime.datetime.now()-state['start']).total_seconds())
    return jsonify({'is_recording':state.get('is_recording',False),
                    'filename':state.get('filename',''), 'elapsed':elapsed})

@app.route('/api/recording/download/<filename>')
def download_recording(filename):
    safe = os.path.basename(filename)
    fp = os.path.join(RECORDING_FOLDER, safe)
    if not os.path.exists(fp):
        return jsonify({'error':'File tidak ditemukan'}), 404
    return send_file(fp, as_attachment=True, download_name=safe, mimetype='video/mp4')

@app.route('/api/recording/list')
def list_recordings():
    files = []
    for f in sorted(os.listdir(RECORDING_FOLDER), reverse=True):
        if f.endswith('.mp4'):
            fp = os.path.join(RECORDING_FOLDER, f)
            cam_label = 'Kamera 1' if f.startswith('cam1_') else 'Kamera 2' if f.startswith('cam2_') else '-'
            files.append({'filename':f,'camera':cam_label,
                          'size_mb':round(os.path.getsize(fp)/(1024*1024),2),
                          'download_url':f'/api/recording/download/{f}'})
    return jsonify({'recordings': files})


if __name__ == '__main__':
    print("=" * 70)
    print("Sistem Deteksi Penyakit Ayam Broiler - 2 KAMERA")
    print("=" * 70)
    print(f"Mode Operasi  : {OPERATION_MODE.upper()}")
    print(f"Confidence    : {CONFIDENCE_THRESHOLD*100:.0f}%")
    print("=" * 70)

    for cid in CAMERA_CONFIG:
        connect_camera(cid)
    if model is None:
        print("PERINGATAN: Model tidak berhasil dimuat")

    app.run(debug=False, host='0.0.0.0', port=5000, threaded=True)